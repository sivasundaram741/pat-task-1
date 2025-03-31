import openpyxl
from datetime import datetime

def write_test_result(file_path, test_id, result):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == test_id:
            sheet.cell(row=row, column=7, value=result)
            sheet.cell(row=row, column=5, value=datetime.now().strftime("%H:%M:%S"))
            break

    workbook.save(file_path)
